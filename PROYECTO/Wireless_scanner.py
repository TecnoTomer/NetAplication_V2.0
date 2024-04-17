import pywifi #modulo para funciones de resdes wifi
import os #modulo para interactuar con el sistema windows
import pandas as pd #modulo para analizar datos
from openpyxl import load_workbook, Workbook #modulo para interactuar con archivos excel
from openpyxl.styles import Alignment, Font, Color, PatternFill #modulo para interactuar con archivos excel
import openpyxl #modulo para interactuar con archivos excel
from  tkinter import ttk, filedialog #modulo para la creacion de ventanas y frames widgets
from tkinter import font as tkFont #modulo para la personalizacion widgets
import time #modulo para funciones de tiempo
import pyperclip
import shutil
import json

import Variables #usar funciones y varibles del script Variables.py
import Alertas #usar funciones y varibles del script Alertas.py
import Wireless_funciones #usar funciones y varibles del script Wireless_funciones.py
from Logs import agregar_log

security_dict = {
    0: "FREE",
    1: "WEP",
    2: "WPA",
    3: "WPA2",
    4: "WPA3",
}

def get_band_and_channel_from_frequency(frequency):
    if 2412 <= frequency <= 2484:
        channel = (frequency - 2407) // 5
        return "2.4 GHz", int(channel)
    elif 5170 <= frequency <= 5825:
        channel = (frequency - 5000) // 5
        return "5 GHz", int(channel)
    else:
        return "N/A", "N/A"

def copiar_archivo(origen, destino): #sin uso aun!
    try:
        shutil.copy(origen, destino)
        Msg = f"se copio archivo de {origen} a {destino}"
        agregar_log(Msg)
    except Exception as e:
        Msg = f"Intento copiar data pero se encontro error: {e}"
        agregar_log(Msg)
        pass

def guardar_en_excel(datos, ruta_guardado):
    try:
        # Verificar si el archivo ya existe
        if os.path.exists(ruta_guardado):
            # Eliminar el archivo existente
            os.remove(ruta_guardado)
            Msg = f"El archivo {ruta_guardado} ya existe y ha sido eliminado."
            agregar_log(Msg)

        time.sleep(2)
        # Crear un nuevo libro de trabajo
        nuevo_libro = openpyxl.Workbook()
        nueva_hoja = nuevo_libro.active

        # Escribir los datos en el nuevo libro, comenzando desde la fila 2
        for fila in datos:
            nueva_hoja.append(fila)

        # Guardar el nuevo libro en la ruta
        nuevo_libro.save(ruta_guardado)
        Msg = f"Los datos han sido guardados en {ruta_guardado}."
        agregar_log(Msg)
    except Exception as e:
        Msg = f"Error al intentar guardar datos en Excel: {e}"
        agregar_log(Msg)
        Alertas.alerta_error(Variables.titulo, Variables.alerta_error, Msg)

def obtener_interfaces_wifi():
    Msg = "Se inicio escaneo de interfaces wireless"
    agregar_log(Msg)
    try:
        wifi = pywifi.PyWiFi()
        interfaces = wifi.interfaces()
        
        if not interfaces:
            Msg = "No se encontraron interfaces Wi-Fi."
            return [], Alertas.alerta_error(Variables.titulo, Variables.alerta_error, Msg)

        # Lista de diccionarios con información de las interfaces
        info_interfaces = []
        for iface in interfaces:
            info_interface = {
                "nombre": iface.name(),
            }
            info_interfaces.append(info_interface)

        return info_interfaces, None
    except Exception as e:
        Msg = f"Error al obtener interfaces Wi-Fi: {e}"
        return [], Alertas.alerta_error(Variables.titulo, Variables.alerta_error, Msg)
        Msg = f"Error al obtener interfaces Wi-Fi: {e}"
        agregar_log(Msg)

def actualizar_combobox():
    interfaces, error = obtener_interfaces_wifi()
    if error:
        Msg = f"Error: {error}"
        Alertas.alerta_error(Variables.titulo, Variables.alerta_error, Msg)
        Msg = f"Error: {error}"
        agregar_log(Msg)
        return

    # Obtener solo los nombres de las interfaces como una tupla
    nombres_interfaces = tuple(iface["nombre"] for iface in interfaces)

    # Actualizar los valores del ComboBox con los nombres de las interfaces
    Wireless_funciones.combo_int ['values'] = nombres_interfaces

def obtener_indice_int(combo_interfaz):
    # Obtener el índice seleccionado desde el ComboBox
    iface_seleccionada_index = combo_interfaz.current()
    Msg = f"se encontro un total de {iface_seleccionada_index} interfaces"
    agregar_log(Msg)
    return iface_seleccionada_index

def scan_wifi(inter):
    Msg = f"Se inicio escaneo de redes wifi por la interface {inter}"
    agregar_log(Msg)
    wifi = pywifi.PyWiFi()
    iface = wifi.interfaces()[inter]
    tiempo_inicio = time.time()  # Obtener el tiempo de inicio
    duracion_escaneo_segundos = Variables.generar_numero_aleatorio(8, 30)
    wifi_list = []  # Agregamos esta línea para almacenar los datos del escaneo
    bssids_seen = set()  # Almacenar las BSSID para evitar duplicados

    while (time.time() - tiempo_inicio) < duracion_escaneo_segundos:
        try:
            iface.scan() #Obtener resultados del escaneo:
            networks = iface.scan_results() #Obtener resultados del escaneo:

            for network in networks: #Procesar cada red WiFi detectada:
                bssid = network.bssid # información esencial de la red:
                if bssid not in bssids_seen: #Evitar duplicados usando BSSID:
                    wifi_name = network.ssid if network.ssid else "'HIDDEN'" #Obtener nombre de la red WiFi (SSID):
                    signal_strength = network.signal #btener fuerza de señal:
                    channel = network.channel if hasattr(network, "channel") else "N/A" #canal de la red WiFi:
                    frequency = network.freq / 1000 if hasattr(network, "freq") else "N/A" #frecuencia de la red WiFi:
                    band, channel = get_band_and_channel_from_frequency(frequency) #obtiene la banda con base a la frecuencia
                    security = network.akm[0] if network.akm else "N/A" #obtengo el tipo de seguridad de la red
                    security_description = security_dict.get(security, "N/A") #le asigno un tipo de seguridad con base al diccionario security_dict
                    wifi_list.append((wifi_name, bssid, signal_strength, channel, frequency, band, security_description)) #creo una lista y agrego toda la informacion
                    bssids_seen.add(bssid) #Agrego BSSID a conjunto para evitar duplicados:
            Msg = f"Se obtuvo informacion de cada red wifi detectada"
            agregar_log(Msg)
        except Exception as e:
            # Manejar la excepción si ocurre algún error durante el escaneo
            Msg = f"Error al escanear las redes Wi-Fi, es posible que el driver de la interface esté desactivado, actívelo y vuelva a intentar: {e}"
            Opcion = Alertas.alerta_error(Variables.titulo, Variables.alerta_error, Msg)
            Msg = f"Se obtuvo un erro durante el escaneo de redes wifi {e}"
            agregar_log(Msg)
            if Opcion == 3:
                pass
            elif Opcion == 4:
                Msg = f"Se reintento el escaneo wifi"
                agregar_log(Msg)
                Wireless_funciones.B_Iniciar_Escaneo() #reintentar escaneo
            elif Opcion == 5:
                pass
            else:
                pass
            return []

    if not os.path.exists(Variables.ruta):
        os.makedirs(Variables.ruta)

    try: #Se crea el archivo y se agregan los datos del escaneo
        #Verificar si el archivo ya existe
        if os.path.exists(Variables.Excel_wifi_lleno):
            #Si existe, eliminarlo
            os.remove(Variables.Excel_wifi_lleno)
            Msg = f"archivo con data ya existia se eliminio y reemplazo"
            agregar_log(Msg)
        time.sleep(2)
        #Crear un nuevo libro de trabajo
        libro = openpyxl.Workbook()

        #Seleccionar la hoja activa
        hoja = libro.active
        
        #Agregar datos a la hoja
        for fila in wifi_list:
            hoja.append(fila)



        #Guardar el libro de trabajo en el archivo .xlsx
        libro.save(Variables.Excel_wifi_lleno)

        libro.close()
    except Exception as e:
        Msg = f"Error al intentar crear el archivo: {e}"
        Opcion = Alertas.alerta_error(Variables.titulo, Variables.alerta_error, Msg)
        agregar_log(Msg)
        if Opcion == 3:
            pass
        elif Opcion == 4:
            Msg = "Se reintento el escaneo"
            agregar_log(Msg)
            Wireless_funciones.B_Iniciar_Escaneo() #reintentar escaneo
        elif Opcion == 5:
            pass
        else:
            pass
        return []

    #Texto centrado al archivo
    book = load_workbook(Variables.Excel_wifi_lleno) #Cargar el libro de trabajo

    sheet = book.active #Seleccionar la hoja activa

    alignment = Alignment(horizontal='center')

    font = Font(color=Color(rgb="000000")) 
    for row in sheet.iter_rows(): 
        for cell in row:
            cell.alignment = alignment 

            cell.font = font

    # Guardar el libro de trabajo
    book.save(Variables.Excel_wifi_lleno)

    Msg = f"El escaneo se completó, presione 'Actualizar OutPut' para visualizar la información"
    Alertas.alerta_ok(Variables.titulo, Variables.alerta_aviso, Msg)
    agregar_log("Escaneo completado")

    return wifi_list

def crear_treeview(frame):
    global columnas
    columnas = ["NOMBRE", "BSSID", "SEÑAL", "CANAL", "FRECUENCIA (GHz)", "BANDA", "SEGURIDAD"]
    treeview = ttk.Treeview(frame, columns=columnas, show="headings", selectmode="extended")

    # Configurar encabezados de columnas
    for col in columnas:
        treeview.heading(col, text=col, anchor="center")
        treeview.column(col, anchor="center")

    # Agregar barra de desplazamiento vertical
    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=treeview.yview)
    treeview.configure(yscrollcommand=scrollbar.set)

    # Colocar el Treeview y la barra de desplazamiento en el frame
    treeview.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")

    # Ajustar el tamaño de las columnas al contenido
    for col in columnas:
        treeview.column(col, width=tkFont.Font().measure(col))

    # Configurar el comportamiento de la barra de desplazamiento
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    Msg = "Se creo el treeview para visualizar la información"
    agregar_log(Msg)

    return treeview

def copiar_seleccion_al_portapapeles(selft):
    # Obtener elementos seleccionados
    elementos_seleccionados = selft.selection()

    if not elementos_seleccionados:
        Msg = "No hay elementos seleccionados."
        Alertas.alerta_Amarilla(Variables.titulo, Variables.alerta_error, Msg)
        return

    datos_a_copiar = []
    for elemento in elementos_seleccionados:
        valores_elemento = selft.item(elemento, "values")
        datos_a_copiar.append("\t".join(map(str, valores_elemento)))

    datos_copiados = "\n".join(datos_a_copiar)

    pyperclip.copy(datos_copiados)
    Msg = "Datos copiados al portapapeles."
    Alertas.alerta_ok(Variables.titulo, Variables.alerta_aviso, Msg)

def seleccionar_todos_los_elementos(selft):
    # Verificar si hay elementos en el Treeview
    if not selft.get_children():
        Msg = "No hay elementos para seleccionar."
        Alertas.alerta_Amarilla(Variables.titulo, Variables.alerta_error, Msg)
        return

    # Seleccionar todos los elementos
    selft.selection_set(selft.get_children())

    # Actualizar la vista del Treeview
    selft.update()

def actualizar_datos(treeview, nombre_archivo):
    # Limpiar datos actuales en el Treeview
    for item in treeview.get_children():
        treeview.delete(item)

    try:
        # Cargar el libro de trabajo
        book = load_workbook(nombre_archivo)
        sheet = book.active

        # Obtener datos de las columnas
        for row in sheet.iter_rows(min_row=1, values_only=True):  # Iniciar desde la segunda fila
            treeview.insert("", "end", values=row)

        # Ajustar el tamaño de las columnas al contenido
        for idx, col_name in enumerate(columnas, start=1):
            treeview.column(col_name, width=tkFont.Font().measure(col_name))
        Msg = "Se actualizaron los datos al treeview"
        agregar_log(Msg)
            
        # Mostrar mensaje cuando el archivo está vacío
        if not os.path.exists(nombre_archivo) or sheet.max_row == 0:
            Msg = "No hay datos que mostrar, realiza el primer escaneo y vuelve a intentarlo."
            Alertas.alerta_aceptar(Variables.titulo, Variables.alerta_aviso, Msg)
            Msg = "No se encontraron datos por mostrar en el output"
            agregar_log(Msg)
    except Exception as e:
        # Otro tipo de error, imprimir en consola
        Msg = "No hay datos que mostrar, realiza el primer escaneo y vuelve a intentarlo."
        Alertas.alerta_Amarilla(Variables.titulo, Variables.alerta_aviso, Msg)
        Msg = f"Ocurrio un error {e}"
        agregar_log(Msg)

def limpiar_datos(treeview):
    # Verificar si el Treeview tiene elementos
    if treeview.get_children():
        # Limpiar datos en el Treeview
        for item in treeview.get_children():
            treeview.delete(item)
            Msg = "Se limpiaron los datos al treeview"
            agregar_log(Msg)
    else:
        Msg = "No hay datos en el OutPut"
        Alertas.alerta_ok(Variables.titulo, Variables.alerta_aviso, Msg)
        agregar_log(Msg)
        pass

def exportar_a_excel(datos_existente, nombre_archivo):
    Msg = "Se inicio la expoertacion de datos"
    agregar_log(Msg)
    try:
        # Crear un nuevo libro de trabajo
        libro_nuevo = openpyxl.Workbook()
        hoja_nueva = libro_nuevo.active

        # Agregar nuevas columnas con centrado, color de fondo y color de texto blanco
        for col_idx, col_name in enumerate(columnas, start=1):
            celda = hoja_nueva.cell(row=1, column=col_idx, value=col_name)
            celda.font = Font(bold=True, color="FFFFFF")  # Texto blanco
            celda.alignment = Alignment(horizontal='center')  # Centrar texto
            celda.fill = PatternFill(start_color="07429D", end_color="07429D", fill_type="solid")  # Fondo azul

        # Copiar datos desde el archivo existente
        libro_existente = openpyxl.load_workbook(datos_existente)
        hoja_existente = libro_existente.active

        for row_idx, row in enumerate(hoja_existente.iter_rows(values_only=True), start=2):
            for col_idx, value in enumerate(row, start=1):
                hoja_nueva.cell(row=row_idx, column=col_idx, value=value)

        # Ajustar el tamaño de las celdas al mayor tamaño de texto más 2 extra
        for col_idx, _ in enumerate(columnas, start=1):
            max_length = 0
            for row in hoja_nueva.iter_rows(min_row=1, max_row=hoja_nueva.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
            adjusted_width = (max_length + 2) * 1.2
            hoja_nueva.column_dimensions[hoja_nueva.cell(row=1, column=col_idx).column_letter].width = adjusted_width
        time.sleep(Variables.Tiempo)
        # Guardar el libro de trabajo
        libro_nuevo.save(nombre_archivo)
        libro_nuevo.close()
        Msg = f"Exportación exitosa a {nombre_archivo}"
        Alertas.alerta_ok(Variables.titulo, Variables.alerta_aviso, Msg)
        agregar_log(Msg)
    except Exception as e:
        Msg = f"Error al intentar exportar datos, realize primero un escaneo:"
        Alertas.alerta_Amarilla(Variables.titulo, Variables.alerta_error, Msg)
        Msg = f"Error al intentar exportar datos {e}"
        agregar_log(Msg)

def importar_desde_excel(treeview):
    Msg = "Se inicio la importacion de datos"
    agregar_log(Msg)
    try:
        # Abrir el explorador de archivos para seleccionar el archivo .xlsx
        archivo_excel = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])

        # Cargar el libro de trabajo
        libro_existente = openpyxl.load_workbook(archivo_excel)
        hoja_existente = libro_existente.active

        # Limpiar datos actuales del Treeview
        for item in treeview.get_children():
            treeview.delete(item)

        #almacenar los datos
        global datos
        datos = []

        # Agregar datos al Treeview, omitiendo la primera fila con los nombres de las columnas
        for fila in hoja_existente.iter_rows(min_row=2, values_only=True):
            treeview.insert("", "end", values=fila)
            datos.append(fila)

        if not os.path.exists(Variables.ruta):
            os.makedirs(Variables.ruta)

        # Guardar datos en otro archivo Excel
        guardar_en_excel(datos, Variables.Excel_wifi_lleno)

        Msg = "Importación exitosa"
        Alertas.alerta_ok(Variables.titulo, Variables.alerta_aviso, Msg)
        agregar_log(Msg)
        time.sleep(Variables.Tiempo)
    except Exception as e:
        Msg = f"Error al intentar importar datos desde Excel: {e}"
        Alertas.alerta_Amarilla(Variables.titulo, Variables.alerta_error, Msg)
        agregar_log(Msg)

def ordenar_por_columna(treeview, columna):
    # Verificar si el Treeview tiene elementos
    if treeview.get_children():
        # Obtener los elementos del Treeview y aplicar filtro
        elementos = [(treeview.item(item, "values"), item) for item in treeview.get_children()]

        # Ordenar los elementos por el valor de la columna seleccionada
        elementos.sort(key=lambda x: x[0][columna] if x[0] and x[0][columna] else 0)

        # Limpiar el Treeview
        treeview.delete(*treeview.get_children())

        # Insertar los elementos ordenados en el Treeview
        for _, item in elementos:
            # Insertar los elementos en las columnas
            treeview.insert("", "end", values=_)
    else:
        Msg = "No hay datos en el OutPut"
        Alertas.alerta_ok(Variables.titulo, Variables.alerta_aviso, Msg)
        pass

def ordenar_por_intensidad(treeview):
    # Verificar si el Treeview tiene elementos
    if treeview.get_children():
        elementos = [(treeview.item(item, "values"), item) for item in treeview.get_children()]

        elementos.sort(key=lambda x: int(x[0][2]) if x[0] and x[0][2] else 0)

        treeview.delete(*treeview.get_children())

        for _, item in elementos:
            treeview.insert("", "end", values=_)
    else:
        Msg = "No hay datos en el OutPut"
        Alertas.alerta_ok(Variables.titulo, Variables.alerta_aviso, Msg)
        pass

def aplicar_filtro_seguridad(treeview, seguridad_seleccionada):
    # Obtener todos los elementos del Treeview
    elementos = [(treeview.item(item, "values"), item) for item in treeview.get_children()]

    # Limpiar el Treeview
    treeview.delete(*treeview.get_children())

    # Recorrer los elementos originales y agregar solo aquellos que cumplan con el filtro de seguridad
    for _, item in elementos:
        original_elemento = _

        # Verificar si la columna de seguridad cumple con el filtro
        if len(original_elemento) > 6 and original_elemento[6] == seguridad_seleccionada:
            # Si cumple, agregarlo al Treeview
            treeview.insert("", "end", values=original_elemento)

def aplicar_filtro_banda(treeview, var_canal_2, var_canal_5):
    # Obtener todos los elementos del Treeview
    elementos = [(treeview.item(item, "values"), item) for item in treeview.get_children()]

    # Limpiar el Treeview
    treeview.delete(*treeview.get_children())

    # Recorrer los elementos originales y agregar solo aquellos que cumplan con el filtro de banda
    for _, item in elementos:
        original_elemento = _

        # Verificar si cumple con el filtro de banda
        if len(original_elemento) > 5 and (
            (var_canal_2.get() and "2.4" in original_elemento[5]) or
            (var_canal_5.get() and "5" in original_elemento[5])
        ):
            # Si cumple, agregarlo al Treeview
            treeview.insert("", "end", values=original_elemento)

def aplicar_filtro_canal(treeview, canal_seleccionado):
    # Obtener todos los elementos del Treeview
    elementos = [(treeview.item(item, "values"), item) for item in treeview.get_children()]

    # Limpiar el Treeview
    treeview.delete(*treeview.get_children())

    # Recorrer los elementos originales y agregar solo aquellos que cumplan con el filtro de canal
    for _, item in elementos:
        original_elemento = _

        # Verificar si la columna de canal cumple con el filtro
        if len(original_elemento) > 3 and original_elemento[3] == canal_seleccionado:
            # Si cumple, agregarlo al Treeview
            treeview.insert("", "end", values=original_elemento)

def recuento(etiqueta, treeview):
    # Obtener el recuento de elementos en la columna 0 del Treeview
    Variables.treeview_recuento = len(treeview.get_children())

    texto = f"Redes Filtradas: {Variables.treeview_recuento}"
    etiqueta.config(text=texto)

    # Establecer el tiempo de espera antes de la próxima actualización (en milisegundos)
    tiempo_espera = 2000  # 5000 milisegundos = 5 segundos
    etiqueta.after(tiempo_espera, lambda: recuento(etiqueta, treeview))

def recuento_escaneados(etiqueta):
    # Verificar si el archivo Excel existe
    if os.path.exists(Variables.Excel_wifi_lleno):
        # Cargar el archivo Excel en un DataFrame
        df = pd.read_excel(Variables.Excel_wifi_lleno, header=None)

        # Verificar si hay al menos una fila en el DataFrame
        if not df.empty:
            # Contar celdas ocupadas en la columna A1
            Variables.v_recuento = df[0].count()
        else:
            Variables.v_recuento = 0
    else:
        Variables.v_recuento = 0

    texto = f"Redes Escaneadas: {Variables.v_recuento}"
    etiqueta.config(text=texto)

    tiempo_espera = 2000  # tiempo en actualizar los datos
    etiqueta.after(tiempo_espera, lambda: recuento_escaneados(etiqueta))

def obtener_seleccion(treeview, etiqueta):
    # Obtener los elementos seleccionados en el Treeview
    elementos_seleccionados = treeview.selection()
    # Obtener el número de elementos seleccionados
    Variables.selecion_recuento = len(elementos_seleccionados)

    texto = f"Celdas seleccionadas: {Variables.selecion_recuento}"
    etiqueta.config(text=texto)
    tiempo_espera = 2000  # tiempo en actualizar los datos
    etiqueta.after(tiempo_espera, lambda: obtener_seleccion(treeview, etiqueta))


def guardar_datos_en_json(treeview):
    #toma los datos del treeview y los guarda en un json
    nombre_archivo = "lib/networks.json"
    # Obtener todas las columnas del Treeview
    columnas = treeview['columns']

    # Obtener todos los datos del Treeview
    datos = []
    for item in treeview.get_children():
        fila = []
        for col in columnas:
            fila.append(treeview.item(item, 'values')[columnas.index(col)])
        datos.append(fila)

    # Crear un diccionario donde cada clave es el nombre de la columna
    # y los valores son las listas de datos correspondientes
    datos_dict = {}
    for idx, col in enumerate(columnas):
        datos_dict[col] = [fila[idx] for fila in datos]

    # Guardar los datos en un archivo JSON
    with open(nombre_archivo, 'w') as json_file:
        json.dump(datos_dict, json_file, indent=4)

    Msg = "Se actualizo la lista de datos de networks a graficar"
    agregar_log(Msg)